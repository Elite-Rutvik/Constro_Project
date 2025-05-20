import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkinter.scrolledtext import ScrolledText
import json
import threading
from web.demo_last_saved import (
    Casting, Shape, optimize_panels, load_castings_from_json, print_results, STANDARD_PANEL_SIZES
)
import sys
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime

class PanelOptimizerUI:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Panel Optimizer System")
        self.root.geometry("1000x800")
        
        self.castings = []
        self.optimization_complete = False
        self.primary_idx = None
        self.setup_ui()

    def setup_ui(self):
        # Create notebook for tabs
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(expand=True, fill='both', padx=10, pady=5)

        # Create main tabs
        self.input_tab = ttk.Frame(self.notebook)
        self.results_tab = ttk.Frame(self.notebook)
        
        self.notebook.add(self.input_tab, text="Input Data")
        self.notebook.add(self.results_tab, text="Results")

        self.setup_input_tab()
        self.setup_results_tab()

    def setup_input_tab(self):
        # Input method selection
        input_frame = ttk.LabelFrame(self.input_tab, text="Data Input Method")
        input_frame.pack(fill='x', padx=10, pady=5)

        self.input_method = tk.StringVar(value="json")
        ttk.Radiobutton(input_frame, text="Load from JSON", variable=self.input_method, 
                       value="json", command=self.toggle_input_method).pack(side='left', padx=5)
        ttk.Radiobutton(input_frame, text="Manual Input", variable=self.input_method, 
                       value="manual", command=self.toggle_input_method).pack(side='left', padx=5)

        # JSON input section
        self.json_frame = ttk.LabelFrame(self.input_tab, text="JSON Input")
        self.json_frame.pack(fill='x', padx=10, pady=5)

        ttk.Button(self.json_frame, text="Browse JSON File", 
                  command=self.browse_json).pack(padx=5, pady=5)
        self.json_path_var = tk.StringVar()
        ttk.Entry(self.json_frame, textvariable=self.json_path_var, 
                 state='readonly').pack(fill='x', padx=5, pady=5)

        # Manual input section
        self.manual_frame = ttk.LabelFrame(self.input_tab, text="Manual Input")
        self.manual_frame.pack(fill='x', padx=10, pady=5)
        
        # Casting controls
        casting_controls = ttk.Frame(self.manual_frame)
        casting_controls.pack(fill='x', padx=5, pady=5)
        
        ttk.Label(casting_controls, text="Casting Name:").pack(side='left', padx=5)
        self.casting_name = ttk.Entry(casting_controls, width=20)
        self.casting_name.pack(side='left', padx=5)
        
        ttk.Button(casting_controls, text="Add Casting", 
                  command=self.add_casting).pack(side='left', padx=5)

        # Shape controls
        shape_controls = ttk.Frame(self.manual_frame)
        shape_controls.pack(fill='x', padx=5, pady=5)
        
        ttk.Label(shape_controls, text="Shape Name:").pack(side='left', padx=5)
        self.shape_name = ttk.Entry(shape_controls, width=20)
        self.shape_name.pack(side='left', padx=5)
        
        ttk.Label(shape_controls, text="Side Lengths:").pack(side='left', padx=5)
        self.side_lengths = ttk.Entry(shape_controls, width=30)
        self.side_lengths.pack(side='left', padx=5)
        
        ttk.Button(shape_controls, text="Add Shape", 
                  command=self.add_shape).pack(side='left', padx=5)

        # Data preview
        preview_frame = ttk.LabelFrame(self.input_tab, text="Data Preview")
        preview_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        self.preview_text = ScrolledText(preview_frame, height=10)
        self.preview_text.pack(fill='both', expand=True, padx=5, pady=5)

        # Optimization controls
        control_frame = ttk.Frame(self.input_tab)
        control_frame.pack(fill='x', padx=10, pady=5)
        
        ttk.Label(control_frame, text="Primary Casting:").pack(side='left', padx=5)
        self.primary_casting_var = tk.StringVar()
        self.primary_casting_select = ttk.Combobox(control_frame, 
                                                  textvariable=self.primary_casting_var)
        self.primary_casting_select.pack(side='left', padx=5)
        
        ttk.Button(control_frame, text="Run Optimization", 
                  command=self.run_optimization).pack(side='right', padx=5)

        # Initially hide manual input
        self.manual_frame.pack_forget()

    def setup_results_tab(self):
        # Results display
        self.results_text = ScrolledText(self.results_tab)
        self.results_text.pack(fill='both', expand=True, padx=10, pady=5)

        # Export buttons frame
        export_frame = ttk.Frame(self.results_tab)
        export_frame.pack(pady=5)
        
        ttk.Button(export_frame, text="Export to Text", 
                  command=self.export_results_text).pack(side='left', padx=5)
        ttk.Button(export_frame, text="Export to Excel", 
                  command=self.export_results_excel).pack(side='left', padx=5)

    def toggle_input_method(self):
        if self.input_method.get() == "json":
            self.manual_frame.pack_forget()
            self.json_frame.pack(after=self.input_frame)
        else:
            self.json_frame.pack_forget()
            self.manual_frame.pack(after=self.input_frame)

    def browse_json(self):
        filename = filedialog.askopenfilename(
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )
        if filename:
            self.json_path_var.set(filename)
            self.load_json_data()

    def load_json_data(self):
        try:
            self.castings = load_castings_from_json(self.json_path_var.get())
            self.update_preview()
            self.update_primary_casting_options()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load JSON: {str(e)}")

    def add_casting(self):
        name = self.casting_name.get().strip()
        if name:
            casting = Casting(name)
            self.castings.append(casting)
            self.casting_name.delete(0, tk.END)
            self.update_preview()
            self.update_primary_casting_options()

    def add_shape(self):
        if not self.castings:
            messagebox.showwarning("Warning", "Please add a casting first")
            return

        name = self.shape_name.get().strip()
        lengths = self.side_lengths.get().strip()
        
        try:
            sides = [int(x) for x in lengths.split(',')]
            shape = Shape(name, sides)
            self.castings[-1].add_shape(shape)
            
            self.shape_name.delete(0, tk.END)
            self.side_lengths.delete(0, tk.END)
            self.update_preview()
        except ValueError:
            messagebox.showerror("Error", "Invalid side lengths. Use comma-separated numbers")

    def update_preview(self):
        self.preview_text.delete('1.0', tk.END)
        for casting in self.castings:
            self.preview_text.insert(tk.END, f"Casting: {casting.name}\n")
            for shape in casting.shapes:
                self.preview_text.insert(tk.END, 
                    f"  Shape: {shape.name}, Sides: {shape.sides}\n")
            self.preview_text.insert(tk.END, "\n")

    def update_primary_casting_options(self):
        options = [casting.name for casting in self.castings]
        self.primary_casting_select['values'] = options
        if options:
            self.primary_casting_select.set(options[0])

    def run_optimization(self):
        if not self.castings:
            messagebox.showwarning("Warning", "No castings available")
            return

        try:
            # Redirect stdout to capture print outputs
            output = io.StringIO()
            sys.stdout = output

            # Find primary casting index
            primary_name = self.primary_casting_var.get()
            self.primary_idx = next(i for i, c in enumerate(self.castings) 
                             if c.name == primary_name)

            # Run optimization in a separate thread
            def optimize():
                try:
                    optimize_panels(self.castings, self.primary_idx)
                    print_results(self.castings, self.primary_idx)
                    
                    # Mark optimization as complete
                    self.optimization_complete = True
                    
                    # Update UI in main thread
                    self.root.after(0, lambda: self.show_results(output.getvalue()))
                except Exception as e:
                    self.root.after(0, lambda: messagebox.showerror("Error", 
                                  f"Optimization failed: {str(e)}"))
                finally:
                    sys.stdout = sys.__stdout__

            threading.Thread(target=optimize, daemon=True).start()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to start optimization: {str(e)}")
            sys.stdout = sys.__stdout__

    def show_results(self, results):
        self.results_text.delete('1.0', tk.END)
        self.results_text.insert('1.0', results)
        self.notebook.select(self.results_tab)

    def export_results_text(self):
        """Export results to text file (original functionality)"""
        filename = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
        )
        if filename:
            with open(filename, 'w') as f:
                f.write(self.results_text.get('1.0', tk.END))

    def export_results_excel(self):
        """Export results to Excel file with two sheets"""
        if not self.optimization_complete:
            messagebox.showwarning("Warning", "Please run optimization first")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if filename:
            try:
                self.create_excel_export(filename)
                messagebox.showinfo("Success", f"Results exported to {filename}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to export to Excel: {str(e)}")

    def create_excel_export(self, filename):
        """Create Excel file with two sheets as specified"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create first sheet: Casting-wise dimensions and panel layouts
        sheet1 = wb.create_sheet("Casting Dimensions & Panels")
        self.create_dimensions_sheet(sheet1)
        
        # Create second sheet: Overall panel summary
        sheet2 = wb.create_sheet("Panel Summary")
        self.create_panel_summary_sheet(sheet2)
        
        # Save the workbook
        wb.save(filename)

    def create_dimensions_sheet(self, sheet):
        """Create sheet showing casting-wise dimensions and panel layouts"""
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Add title
        sheet.append(["Panel Optimization Results - Casting Dimensions & Panel Layouts"])
        sheet["A1"].font = Font(bold=True, size=14)
        sheet.merge_cells("A1:F1")
        
        # Add timestamp
        sheet.append([f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        sheet.append([])  # Empty row
        
        row_num = 4
        
        for i, casting in enumerate(self.castings):
            # Casting header
            casting_type = "PRIMARY" if i == self.primary_idx else "SECONDARY"
            sheet.append([f"Casting: {casting.name} ({casting_type})"])
            sheet[f"A{row_num}"].font = Font(bold=True, size=12)
            sheet.merge_cells(f"A{row_num}:F{row_num}")
            row_num += 1
            
            # Column headers for this casting
            headers = ["Shape", "Side", "Length (mm)", "Panel Layout", "Panel Count", "Panel Types"]
            sheet.append(headers)
            
            # Apply header styling
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=row_num, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            row_num += 1
            
            # Add shape and side data
            for shape in casting.shapes:
                for side_idx, (side_length, panels) in enumerate(zip(shape.sides, shape.panel_layout)):
                    panel_layout_str = str(panels).replace('[', '').replace(']', '').replace(',', ' +')
                    panel_count = len(panels)
                    
                    # Determine panel types
                    standard_count = sum(1 for p in panels if p in STANDARD_PANEL_SIZES)
                    custom_count = len(panels) - standard_count
                    panel_types = f"Std: {standard_count}, Custom: {custom_count}"
                    
                    sheet.append([
                        shape.name if side_idx == 0 else "",  # Show shape name only on first side
                        f"Side {side_idx + 1}",
                        side_length,
                        panel_layout_str,
                        panel_count,
                        panel_types
                    ])
                    row_num += 1
            
            # Add empty row between castings
            sheet.append([])
            row_num += 1
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

    def create_panel_summary_sheet(self, sheet):
        """Create sheet showing overall panel usage summary"""
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Calculate panel statistics
        all_panels = {}
        standard_panels = {}
        custom_panels = {}
        
        for casting in self.castings:
            for shape in casting.shapes:
                for panels in shape.panel_layout:
                    for panel in panels:
                        all_panels[panel] = all_panels.get(panel, 0) + 1
                        if panel in STANDARD_PANEL_SIZES:
                            standard_panels[panel] = standard_panels.get(panel, 0) + 1
                        else:
                            custom_panels[panel] = custom_panels.get(panel, 0) + 1
        
        # Add title
        sheet.append(["Panel Usage Summary"])
        sheet["A1"].font = Font(bold=True, size=14)
        sheet.merge_cells("A1:C1")
        
        # Add timestamp
        sheet.append([f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        sheet.append([])  # Empty row
        
        row_num = 4
        
        # Standard Panels Section
        sheet.append(["Standard Panels"])
        sheet[f"A{row_num}"].font = Font(bold=True, size=12)
        sheet.merge_cells(f"A{row_num}:C{row_num}")
        row_num += 1
        
        # Headers for standard panels
        headers = ["Panel Size (mm)", "Count", "Type"]
        sheet.append(headers)
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=row_num, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        row_num += 1
        
        # Add standard panel data
        if standard_panels:
            for size in sorted(standard_panels.keys()):
                count = standard_panels[size]
                sheet.append([f"{size}mm", count, "Standard"])
                row_num += 1
        else:
            sheet.append(["No standard panels used", "", ""])
            row_num += 1
        
        # Add empty row
        sheet.append([])
        row_num += 1
        
        # Custom Panels Section
        sheet.append(["Custom Panels"])
        sheet[f"A{row_num}"].font = Font(bold=True, size=12)
        sheet.merge_cells(f"A{row_num}:C{row_num}")
        row_num += 1
        
        # Headers for custom panels
        sheet.append(headers)
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=row_num, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        row_num += 1
        
        # Add custom panel data
        if custom_panels:
            for size in sorted(custom_panels.keys()):
                count = custom_panels[size]
                sheet.append([f"{size}mm", count, "Custom"])
                row_num += 1
        else:
            sheet.append(["No custom panels used", "", ""])
            row_num += 1
        
        # Add empty row
        sheet.append([])
        row_num += 1
        
        # Summary Statistics
        sheet.append(["Summary Statistics"])
        sheet[f"A{row_num}"].font = Font(bold=True, size=12)
        sheet.merge_cells(f"A{row_num}:C{row_num}")
        row_num += 1
        
        total_panels = sum(all_panels.values())
        total_standard = sum(standard_panels.values())
        total_custom = sum(custom_panels.values())
        
        summary_data = [
            ["Total Panels Used", total_panels, ""],
            ["Standard Panels", total_standard, f"{(total_standard/total_panels*100):.1f}%" if total_panels > 0 else "0%"],
            ["Custom Panels", total_custom, f"{(total_custom/total_panels*100):.1f}%" if total_panels > 0 else "0%"],
            ["Standard Panel Types", len(standard_panels), ""],
            ["Custom Panel Types", len(custom_panels), ""],
            ["Total Panel Types", len(all_panels), ""]
        ]
        
        for data_row in summary_data:
            sheet.append(data_row)
            row_num += 1
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    app = PanelOptimizerUI()
    app.run()